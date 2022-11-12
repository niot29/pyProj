'''
Tutorial on handling csv fil
Moduloe dependeci
 - openpyxl
'''
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('Grades.xlsx')

ws = wb.active
print(ws['A2'].value) # Read value from a cell 'A2'
ws['A2'] = "Test" # set a value in cell 'A2' but no sav
wb.save('Grades.xlsx') # Save sheet after add value

print(wb.sheetnames)

# Create a nwe workbook
wb = Workbook()
ws = wb.active
ws.title = "Data"
ws.append(['Time','Is','Great','!'])
wb.save('tim.xlsx')

wb = load_workbook('tim.xlsx')
ws = wb.active

# Loop true cell
for row in range(1,11):
    for col in range(1,5):
        char = get_column_letter(col)
       # print(ws[char + str(row)].value)

# merge cell
ws.merge_cells("A1:D1")

#Un merge cell
ws.unmerge_cells("A1:D1")

# insert row
ws.insert_rows(7)

# Delete row
ws.delete_rows(7)

# insert column
ws.insert_cols(2)

# delte column
ws.delete_cols(2)

# move col
ws.move_range("C1:D11", rows=2, cols=2)

wb.save('tim.xlsx')
